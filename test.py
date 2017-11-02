# # from openpyxl import Workbook
# # wb = Workbook()
# #
# # # grab the active worksheet
# # ws = wb.active
# #
# # # Data can be assigned directly to cells
# # ws['A1'] = 42
#
# from openpyxl import load_workbook
#
# wb = load_workbook(filename = 'test.xlsx')
#
# # sheet_ranges = wb['range names']
#
# print wb
# print wb['A1'].value

import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

print wb

print wb.get_sheet_names()

print sheet
